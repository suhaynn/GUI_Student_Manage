import os
import time
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Alignment
import sql_link


class ExcelWriter:
    def __init__(self):
        self.headers = ['用户名', '学号', '语文成绩', '数学成绩', '英语成绩']
        self.data = []

    def get_data(self):
        # 连接数据库
        db = sql_link.Database()
        # 构造查询语句
        query = "SELECT username, stu_id, chinese_score, math_score, english_score FROM user WHERE user_type = 0 order by  'stu_id'"
        # 执行查询操作
        self.data = db.fetchall(query)
        # 关闭数据库连接
        db.close()
        # 对数据进行去重
        self.data = list(set(self.data))

    def write_excel(self):
        # 创建workbook对象
        wb = openpyxl.Workbook()
        ws = wb.active
        # 设置表头
        ws.append(self.headers)
        # 写入数据
        for row in self.data:
            ws.append(row)
            # 设置单元格格式
            for i in range(len(row)):
                cell = ws.cell(row=ws.max_row, column=i + 1)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 弹出确认窗口
        confirmation = messagebox.askquestion("保存成功", "文件已保存成功，是否打印？")
        if confirmation == 'yes':
            # 保存文件
            timestamp = time.strftime("_%Y-%m-%d_%H-%M-%S", time.localtime(time.time()))
            file_name = "user_data" + timestamp + ".xlsx"
            file_path = os.path.join(".", file_name)
            wb.save(file_path)
            print("打印文件：", file_path)

# if __name__ == '__main__':
#     ew = ExcelWriter()
#     ew.get_data()
#     # 写入Excel文件并打印
#     ew.write_excel()